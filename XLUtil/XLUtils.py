import openpyxl

class excelOps():

    def __init__(self,xlFile,xlSheet):
        self.xlFile = xlFile
        self.xlSheet = xlSheet
        self.xlwb = openpyxl.load_workbook(self.xlFile)
        self.xlws = self.xlwb.get_sheet_by_name(self.xlSheet)

    def getRowCount(self):
        return(self.xlws.max_row)

    def getColCount(self):
        return(self.xlws.max_column)

    def readData(self,rowNum,colNum):
        return (self.xlws.cell(row = rowNum , column = colNum).value)

    def writeData(self,rowNum,colNum,data):
        self.xlws.cell(row = rowNum , column = colNum).value = data
        self.xlwb.save(self.xlFile)
        self.xlwb.close()
